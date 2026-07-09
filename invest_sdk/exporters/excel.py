from __future__ import annotations

from pathlib import Path
from typing import Any


class ExcelExporter:
    """Экспорт данных в Excel (.xlsx).

    Requires: openpyxl
    """

    @staticmethod
    def export(
        data: list[dict[str, Any]],
        file_path: str | Path,
        *,
        sheet_name: str = "Data",
    ) -> Path:
        """Экспортировать список словарей в Excel файл.

        Args:
            data: Данные для экспорта.
            file_path: Путь к файлу.
            sheet_name: Имя листа.

        Returns:
            Путь к созданному файлу.
        """
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
        except ImportError:
            msg = "ExcelExporter requires openpyxl. Install it with: pip install openpyxl"
            raise ImportError(msg) from None

        path = Path(file_path)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        if not data:
            wb.save(str(path))
            return path

        fieldnames = list(data[0].keys())

        for col_idx, field in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=col_idx, value=field)
            cell.font = openpyxl.styles.Font(bold=True)

        for row_idx, row in enumerate(data, 2):
            for col_idx, field in enumerate(fieldnames, 1):
                ws.cell(
                    row=row_idx,
                    column=col_idx,
                    value=str(row.get(field, "")),
                )

        for col_idx in range(1, len(fieldnames) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for row in ws.iter_rows(
                min_col=col_idx,
                max_col=col_idx,
                values_only=True,
            ):
                cell_value = str(row[0]) if row[0] is not None else ""
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            ws.column_dimensions[column_letter].width = min(
                max_length + 2, 50
            )

        wb.save(str(path))
        return path

    @staticmethod
    def export_portfolio(
        portfolio_data: list[dict[str, Any]],
        file_path: str | Path,
    ) -> Path:
        """Экспортировать портфель в Excel."""
        return ExcelExporter.export(
            portfolio_data, file_path, sheet_name="Portfolio"
        )

    @staticmethod
    def export_candles(
        candles_data: list[dict[str, Any]],
        file_path: str | Path,
    ) -> Path:
        """Экспортировать свечи в Excel."""
        return ExcelExporter.export(
            candles_data, file_path, sheet_name="Candles"
        )
